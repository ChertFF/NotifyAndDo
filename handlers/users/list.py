import logging
from time import strftime, gmtime

from aiogram import types
from loader import dp, bot
from utils.db_api.quick_commands import get_task_data
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment
import aiogram.utils.markdown as fmt

LIMIT_TASK = 5  # Максимальное количество задач для каждого статуса


@dp.message_handler(commands=['list'], state="*")
async def cmd_list(message: types.Message):
    chat_id = message.chat.id

    # Получаем все задачи
    all_tasks = await get_task_data(chat_id)

    # Фильтрация задач по статусам и ограничение до LIMIT_TASK
    new_tasks = [task for task in all_tasks if task['Статус'] == 'Новая'][:LIMIT_TASK]
    clarification_tasks = [task for task in all_tasks if task['Статус'] == 'На уточнении'][:LIMIT_TASK]

    # Для closed_tasks добавляем сортировку по updated_at
    closed_tasks = sorted(
        [task for task in all_tasks if task['Статус'] == 'Завершена'],
        key=lambda task: task['Дата закрытия'],
        reverse = True
    )[:LIMIT_TASK]

    # Создаем Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    # Форматирование
    header_font = Font(name='Arial', size=12, bold=True)
    cell_font = Font(name='Arial', size=11)
    cell_alignment = Alignment(wrap_text=False)

    # Заголовок
    headers = [
        "Дата создания", "Дата изменения", "Создал", "Создал (username)", "Создал (fullname)", "Создал (url)",
        "Срок", "Название задачи", "Описание задачи", "Ответственный", "Статус", "Решение", "Дата закрытия",
        "Закрыл", "Закрыл (username)", "Закрыл (fullname)", "Закрыл (url)"
    ]
    ws.append(headers)
    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for cell in col:
            cell.font = header_font
            cell.alignment = cell_alignment

    # Данные
    for task in all_tasks:
        row = [
            task["Дата создания"].replace(tzinfo=None) if task["Дата создания"] else None,
            task["Дата изменения"].replace(tzinfo=None) if task["Дата изменения"] else None,
            task["Создал"], task["Создал (username)"], task["Создал (fullname)"], task["Создал (url)"],
            task["Срок"],
            task["Название задачи"], task["Описание задачи"], task["Ответственный"], task["Статус"],
            task["Решение"],
            task["Дата закрытия"].replace(tzinfo=None) if task["Дата закрытия"] else None,
            task["Закрыл"], task["Закрыл (username)"], task["Закрыл (fullname)"], task["Закрыл (url)"]
        ]
        ws.append(row)
        for col in ws.iter_cols(min_row=ws.max_row, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in col:
                cell.font = cell_font
                cell.alignment = cell_alignment

    # Автоподбор ширины для первой строки
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.4
        ws.column_dimensions[column].width = adjusted_width

    # Скрыть столбцы C, F, N и Q
    hidden_columns = ['C', 'F', 'N', 'Q']
    for col in hidden_columns:
        ws.column_dimensions[col].hidden = True

    # Применить стиль к первой строке для выравнивания по центру
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    time = strftime("%m.%d.%Y_%H.%M.%S", gmtime())
    filename = f'out/task_data_{time}.xlsx'
    wb.save(filename)


    # Определяем текущую дату
    current_date = datetime.now().date()

    # Подсчет просроченных задач
    overdue_tasks_count = sum(1 for task in all_tasks if task["Срок"] < current_date)

    # Подсчет задач, у которых дедлайн истекает менее чем через 7 дней
    close_deadline_tasks_count = sum(1 for task in all_tasks if task["Срок"] - current_date <= timedelta(days=7)) - overdue_tasks_count

    # Формируем текстовое сообщение
    message_text = f"{fmt.hbold('Список задач')}\n\n" \
                   f"{fmt.hbold('📍 Новые:')}\n{format_tasks_list(new_tasks)}\n\n" \
                   f"{fmt.hbold('🔙 На уточнении:')}\n{format_tasks_list(clarification_tasks)}\n\n" \
                   f"{fmt.hbold('🏷 Закрытые: ')}\n{format_tasks_list(closed_tasks)}\n\n" \
                   f"{fmt.hbold('Просрочено:')} {overdue_tasks_count}\n" \
                   f"{fmt.hbold('Близки к дедлайну:')} {close_deadline_tasks_count}"

    # Отправляем текстовое сообщение и Excel файл
    await bot.send_document(chat_id, types.InputFile(filename), caption=message_text)

    logging.warning(
        f'Отправил сообщение /list | {message.chat.id} {message.from_user.id} {message.from_user.full_name} {message.from_user.username}  {message.from_user.url}')

def format_tasks_list(tasks):
    if tasks:
        formatted_tasks = []
        for task in tasks:
            emoji = get_task_emoji(task)
            formatted_tasks.append(f"{emoji} {task['Название задачи']} - {task['Срок'].strftime('%Y-%m-%d')}")
        return "\n".join(formatted_tasks)
    else:
        return f"{fmt.hstrikethrough('Нет задач')}"

def get_task_emoji(task):
    if task["Статус"] == 'Завершена':
        if task["Дата закрытия"] and task["Дата закрытия"].date() <= task["Срок"]:
            return '✅'  # Зеленая галочка
        else:
            return '❌'  # Красная галочка
    else:
        now = date.today()
        if task["Срок"] >= now:
            return '🔘'  # Зеленая точка
        else:
            return '❌'  # Красная точка